#! python3

import openpyxl, os, shutil, re, time, datetime

os.chdir('D:\Dropbox\ebayPythonScripts') # setup current working directory


# Search directory for file with .xlsx extension
for files in (os.listdir('.')):
    if not (files.endswith('.xlsx') and files.startswith('DispatchToday')):
        continue
    excellFile = os.path.basename(files)

print(excellFile + '\n')
    
# TODO: Date function and variable, string variable
dt = datetime.datetime.now()
date = dt.strftime('%d%m%Y-%H%M%S')
print(date + '\n')

# Excell sheet settings
wb = openpyxl.load_workbook(excellFile)
sheetname = wb.sheetnames[0]
sheet = wb[sheetname]
maxRowN = sheet.max_row + 1
print(maxRowN)

## TODO: Add value EMP001 EmptyFile
## Loop through N row and add EMP001 EmptyFile in to the empty cells
##for i in range(2, maxRowN):
##    if sheet['N' + str(i)].value == None:
##        sheet['N' + str(i)] = 'EMP001 EmptyFile'
##
##for i in range(2, maxRowN):
##    if sheet['O' + str(i)].value == None:
##        sheet['O' + str(i)] = '1'
##wb.save('DispatchToday.xlsx')
## NOT USING THIS DUE TO TRY EXCEPT, TO BE REMOVED LATER

filesList = []

# Regex for image code
imageRegex = re.compile(r'(\w)+\d\d\d')               
# Regex for phone model
phoneModelRegex = re.compile(r'Samsung S9\+|Samsung S9|Samsung Note 8|Samsung Note 5|Samsung Note 4|Samsung S8\+|Samsung S8|Samsung S7 Edge|Samsung S7|Samsung S6 Edge\+|Samsung S6 Edge|Samsung S6|Samsung S5|Samsung S4|iPhone X|iPhone 8 Plus|iPhone 8 PLUS|iPhone 8|iPhone 7 Plus|iPhone 7 PLUS|iPhone 7|iPhone 6 Plus|iPhone 6 PLUS|iPhone 6|iPhone 5C|iPhone 5|iPhone 4')

# Create file to which list will be uploaded
printingListFile = open('filesToBePrinted' + str(date) + '.txt', 'w') 
print(printingListFile) #prints new text filename

for i in range(2, maxRowN):
    # TODO: What if someone orders SCREEN PROTECTOR?
    try:
        matchImage = imageRegex.search(sheet['N' + str(i)].value)
        matchPhone = phoneModelRegex.search(sheet['N' + str(i)].value)
        #print(matchImage.group() + ' ' + matchPhone.group())
        file = [matchImage.group() + ' ' + matchPhone.group() + '.jpg']
        filesList.append(file)

    except:
        matchImage = None
        matchPhone = None
        continue
    #print(file)
    
    if int(sheet[('O' + str(i))].value) > 1:
        for j in range(1, int(sheet['O' + str(i)].value)):
            filesList.append(file)
    
for filePathsIndex in range(len(filesList)):
    printingListFile.write(str(filesList[filePathsIndex]).strip('[\'\']')+ '\n')

printingListFile.close()
                        
#Prints content of the list
for item in range (len(filesList)):
    print(str(filesList[item]) + '\n')

# Moving files & renaming excell file
shutil.move('D:\Dropbox\ebayPythonScripts\\' + excellFile, 'D:\Dropbox\ebayPythonScripts\ProcessedExcell\\' + 'Completed ' + date + ' ' + excellFile)

# execute readingCSV
time.sleep(1)

#os.system('printingCSV.py')
#execfile('printingCSV.py')
